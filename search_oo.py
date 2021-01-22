import os, socket
from openpyxl import load_workbook
from ping3 import ping

wb = load_workbook('./filial.xlsx')

sheet = wb.active
rows = sheet.max_row
cols = sheet.max_column

data_base = {}

for i in range(2, rows+1):
    email = sheet.cell(row=i, column=8)
    if email.value:
        if email.value != 'нет':
            kod_oo = email.value.split('@')[0]
            if len(kod_oo) > 9:
                kod_oo = 'gm_' + kod_oo[-6:]
            elif kod_oo[:2] == 'ma':
                kod_oo = 'ap_' + kod_oo[2:]
            else:
                kod_oo = kod_oo[:2] + '_' + kod_oo[2:]
            form = sheet.cell(row=i, column=3).value
            kod = sheet.cell(row=i, cell=4).value
            name = sheet.cell(row=i, column=5).value
            address = sheet.cell(row=i, column=7).value
            email = email.value
            tel_number = sheet.cell(row=i, column=9).value

            data_base[kod_oo] = [form, kod, name, address, email, tel_number]

while True:
    print("По умолчанию поиск ведется по коду магазина и наименованию")
    print("Для поиска по адресу введите перед строкой поиска знак '!' (Пример: !40 лет победы)")
    find_oo = input("Введите строку поиска >> ")
    find_sum = []
    if find_oo[0] =='!':
        for key,value in data_base.items:
            if find_oo[1:].lower() in data_base[key][3].lower():
                find_sum.append(key)
    else:
        for key,value in data_base.items():
            if find_oo in data_base[key][1] or find_oo.lower() in data_base[key][2].lower():
                find_sum.append(key)
    if len(find_sum) == 0:
        print("По вашему запросу совпадении не найдено, уточните параметры поиска")            
        continue
    elif len(find_sum) == 1:
        key = find_sum[0]
        print("По вашему запросу найдено 1 совпадение:")
        print("_______________________________________")
        print(f"{data_base[key][0]} {data_base[key][2]} {data_base[key][1]}")
        print(f"{data_base[key][3]}")
        print(f"{data_base[key][4]} {data_base[key][5]}")
    else:
        print(f"По вашему запросу найдено {len(find_sum)} совпадений:")
        for i in range(len(find_sum)):
            key = find_sum[i]
            print("_______________________________________")
            print(f"{i+1}. {data_base[key][0]} {data_base[key][2]} {data_base[key][1]}")
            print(f"    {data_base[key][3]}")
            print(f"    {data_base[key][4]} {data_base[key][5]}")
        try:
            s = int(input(f"Выберите вариант для дальнейшей работы (число от 1 до {len(find_sum)}) >> "))
            if 0 <= s-1 < len(find_sum):
                key = find_sum[s-1]
            else:
                print("Введенное число вне диапазона выборки, продолжаем работу с последнним найденным вариантом")
        except:
            print("Вы ввели не число, продолжаем работу с последним найденным вариантом")

    print()
    print(f" ==> Доступность обьекта      {data_base[key][2]}")
    print()
    hosts = []
    for i,j,k in [key,'Основной ',0],[key+'_1',"Резервный",1]:
        host = 'o' + i + '.online'
        response = ping(host)
        if response:
            response = True
        try:
            hosts.append([host, socket.gethostbyname(host), response])
        except:
            hosts.append([host, False, response])
        if response:
            if hosts[k][2]:
                print(f"    {j} канал связи   {host[k][1]} ==> Онлайн")
            else:
                print(f"    {j} канал связи   Не удалось получить IP-адрес ==> Онлайн")
        else:
            if hosts[k][2]:
                print(f"    {j} канал связи   {host[k][1]} ==> Оффлайн")
            else:
                print(f"    {j} канал связи   Не удалось получить IP-адрес ==> Оффлайн")
        print()
    print("__________________________________________")
